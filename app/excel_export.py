from __future__ import annotations

import json
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_ai_json_safe(asta) -> dict:
    raw = getattr(asta, "ai_result_json", None)
    if not raw:
        return {}
    try:
        if isinstance(raw, dict):
            return raw
        return json.loads(raw)
    except Exception:
        return {}


def stringify(value) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return ", ".join(str(x).strip() for x in value if str(x).strip())
    if isinstance(value, dict):
        try:
            return json.dumps(value, ensure_ascii=False)
        except Exception:
            return str(value)
    return str(value).strip()


def first_non_empty(*values) -> str:
    for value in values:
        if value is None:
            continue
        if isinstance(value, str):
            v = value.strip()
            if v and v.lower() not in {"null", "none", "n.d.", "nd", "-"}:
                return v
        elif isinstance(value, (list, dict)):
            s = stringify(value)
            if s:
                return s
        else:
            s = str(value).strip()
            if s:
                return s
    return ""


def set_link(cell, url: str, label: str):
    if not url:
        cell.value = ""
        return
    cell.value = label
    cell.hyperlink = url
    cell.style = "Hyperlink"


def auto_fit_columns(ws):
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)

        for cell in col_cells:
            try:
                cell_value = "" if cell.value is None else str(cell.value)
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except Exception:
                pass

        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 60)


def style_header_row(ws, headers: list[str]):
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for c, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)


def style_data_row(ws, row_idx: int, values: list):
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def extract_risk(asta, ai_data: dict | None = None) -> str:
    ai_data = ai_data or {}

    rischio_ai = clean(ai_data.get("rischio_operazione"))
    if rischio_ai:
        return rischio_ai

    abusi = clean(getattr(asta, "abusi", None)).lower()
    occupazione = clean(getattr(asta, "occupazione", None)).lower()
    pregiudizievoli = clean(getattr(asta, "pregiudizievoli", None)).lower()

    flags = 0

    if abusi and abusi not in {"nessuno", "nessuna", "non rilevati", "assenti"}:
        flags += 1

    if occupazione and any(x in occupazione for x in ["occupato", "debitore", "terzi", "senza titolo"]):
        flags += 1

    if pregiudizievoli and pregiudizievoli not in {"nessuna", "nessuno", "assenti"}:
        flags += 1

    if flags >= 3:
        return "Alto"
    if flags == 2:
        return "Medio"
    if flags == 1:
        return "Medio-Basso"
    return "Basso"


def build_excel_export(
    aste: Iterable,
    output_path: str | Path | None = None,
    base_url: str = "http://127.0.0.1:8000",
) -> str:
    """
    Esporta le aste in un file Excel con:
    - Foglio 1: Aste
    - Foglio 2: Analisi approfondita
    """

    wb = Workbook()

    # =========================================================
    # FOGLIO 1 - ASTE
    # =========================================================
    ws1 = wb.active
    ws1.title = "Aste"

    headers_main = [
        "ID",
        "Tribunale",
        "RGE",
        "Lotto",
        "Data asta",
        "Prezzo base",
        "Offerta minima",
        "Valore perizia",
        "Rischio",
        "Città",
        "Indirizzo",
        "Proprietario manuale",
        "Occupazione",
        "Foglio",
        "Mappale",
        "Subalterno",
        "Creditore",
        "Stato pratica",
        "Stato perizia",
        "Stato avviso",
        "Stato AI",
        "URL asta",
        "Scheda",
        "File perizia",
        "File avviso",
        "Note",
        "Note operative",
    ]
    style_header_row(ws1, headers_main)

    row_idx = 2

    for asta in aste:
        ai_data = parse_ai_json_safe(asta)

        tribunale_export = clean(getattr(asta, "tribunale", None))
        citta_export = clean(getattr(asta, "citta", None))
        indirizzo_export = clean(getattr(asta, "indirizzo", None))
        occupazione = first_non_empty(getattr(asta, "occupazione", None), ai_data.get("occupazione"))
        foglio = first_non_empty(getattr(asta, "foglio", None), ai_data.get("foglio"))
        mappale = first_non_empty(getattr(asta, "mappale", None), ai_data.get("mappale"))
        subalterno = first_non_empty(getattr(asta, "subalterno", None), ai_data.get("subalterno"))
        creditore = first_non_empty(getattr(asta, "creditore_procedente", None), ai_data.get("creditore_procedente"))
        prezzo_base = first_non_empty(getattr(asta, "prezzo_base", None), ai_data.get("prezzo_base"))
        offerta_minima = first_non_empty(getattr(asta, "offerta_minima", None), ai_data.get("offerta_minima"))
        valore_perizia = first_non_empty(getattr(asta, "valore_perizia", None), ai_data.get("valore_perizia"))
        rischio = extract_risk(asta, ai_data)

        row_main = [
            getattr(asta, "id", None),
            tribunale_export,
            clean(getattr(asta, "rge", None)),
            clean(getattr(asta, "lotto", None)),
            clean(getattr(asta, "data_asta", None)),
            prezzo_base,
            offerta_minima,
            valore_perizia,
            rischio,
            citta_export,
            indirizzo_export,
            clean(getattr(asta, "proprietario", None)),
            occupazione,
            foglio,
            mappale,
            subalterno,
            creditore,
            clean(getattr(asta, "stato_pratica", None)),
            clean(getattr(asta, "perizia_status", None)),
            clean(getattr(asta, "avviso_status", None)),
            clean(getattr(asta, "ai_status", None)),
            "",
            "",
            clean(getattr(asta, "perizia_file_path", None)),
            clean(getattr(asta, "avviso_file_path", None)),
            clean(getattr(asta, "note", None)),
            clean(getattr(asta, "note_operativi", None)),
        ]

        style_data_row(ws1, row_idx, row_main)

        asta_url = clean(getattr(asta, "url", None))
        scheda_url = f"{base_url}/aste/{getattr(asta, 'id', '')}"

        set_link(ws1.cell(row=row_idx, column=22), asta_url, "Apri annuncio")
        set_link(ws1.cell(row=row_idx, column=23), scheda_url, "Apri scheda")

        row_idx += 1

    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = ws1.dimensions
    auto_fit_columns(ws1)

    # =========================================================
    # FOGLIO 2 - ANALISI APPROFONDITA
    # =========================================================
    ws2 = wb.create_sheet(title="Analisi approfondita")

    headers_analysis = [
        "ID",
        "Tribunale",
        "RGE",
        "Città",
        "Indirizzo",
        "Lotto",
        "Data asta",
        "Valore perizia",
        "Prezzo base",
        "Offerta minima",
        "Descrizione immobile",
        "Stato manutentivo",
        "Catasto",
        "Foglio",
        "Mappale",
        "Subalterno",
        "Proprietario",
        "Tipologia immobile",
        "Superficie",
        "Categoria catastale",
        "Classe catastale",
        "Rendita catastale",
        "Agibilità",
        "Impianti",
        "Vincoli e oneri",
        "Conformità catastale",
        "Stato catastale",
        "Conformità urbanistica",
        "Stato urbanistico",
        "Abusi / difformità",
        "Criticità principali",
        "Occupazione",
        "Stato occupazione dettaglio",
        "Pregiudizievoli",
        "Pregiudizievoli dettaglio",
        "Creditore procedente",
        "Debiti condominiali",
        "Rischio operazione",
        "Punti di attenzione investitore",
        "Costi probabili",
        "Valutazione operativa",
        "Strategia consigliata",
        "Note investitore",
        "Sintesi finale",
        "Stato pratica",
        "Note operative",
        "Scheda",
    ]
    style_header_row(ws2, headers_analysis)

    row_idx = 2

    for asta in aste:
        ai_data = parse_ai_json_safe(asta)

        tribunale = clean(getattr(asta, "tribunale", None))
        rge = clean(getattr(asta, "rge", None))
        citta = clean(getattr(asta, "citta", None))
        indirizzo = clean(getattr(asta, "indirizzo", None))
        lotto = clean(getattr(asta, "lotto", None))
        data_asta = clean(getattr(asta, "data_asta", None))
        valore_perizia = first_non_empty(getattr(asta, "valore_perizia", None), ai_data.get("valore_perizia"))
        prezzo_base = first_non_empty(getattr(asta, "prezzo_base", None), ai_data.get("prezzo_base"))
        offerta_minima = first_non_empty(getattr(asta, "offerta_minima", None), ai_data.get("offerta_minima"))

        descrizione_immobile = first_non_empty(
            getattr(asta, "descrizione_immobile", None),
            ai_data.get("descrizione_immobile"),
        )
        stato_manutentivo = first_non_empty(
            getattr(asta, "stato_manutentivo", None),
            ai_data.get("stato_manutentivo"),
        )
        catasto = first_non_empty(getattr(asta, "catasto", None), ai_data.get("catasto"))
        foglio = first_non_empty(getattr(asta, "foglio", None), ai_data.get("foglio"))
        mappale = first_non_empty(getattr(asta, "mappale", None), ai_data.get("mappale"))
        subalterno = first_non_empty(getattr(asta, "subalterno", None), ai_data.get("subalterno"))
        proprietario = first_non_empty(getattr(asta, "proprietario", None), ai_data.get("proprietario"))

        tipologia_immobile = stringify(ai_data.get("tipologia_immobile"))
        superficie = stringify(ai_data.get("superficie"))
        categoria_catastale = stringify(ai_data.get("categoria_catastale"))
        classe_catastale = stringify(ai_data.get("classe_catastale"))
        rendita_catastale = stringify(ai_data.get("rendita_catastale"))
        agibilita = stringify(ai_data.get("agibilita"))
        impianti = stringify(ai_data.get("impianti"))
        vincoli_oneri = stringify(ai_data.get("vincoli_oneri"))
        conformita_catastale = first_non_empty(
            getattr(asta, "conformita_catastale", None),
            ai_data.get("conformita_catastale"),
        )
        stato_catastale = stringify(ai_data.get("stato_catastale"))
        conformita_urbanistica = first_non_empty(
            getattr(asta, "conformita_urbanistica", None),
            ai_data.get("conformita_urbanistica"),
        )
        stato_urbanistico = stringify(ai_data.get("stato_urbanistico"))
        abusi = first_non_empty(getattr(asta, "abusi", None), ai_data.get("abusi"))
        criticita_principali = stringify(ai_data.get("criticita_principali"))

        occupazione = first_non_empty(getattr(asta, "occupazione", None), ai_data.get("occupazione"))
        stato_occupazione_dettaglio = stringify(ai_data.get("stato_occupazione_dettaglio"))
        pregiudizievoli = first_non_empty(getattr(asta, "pregiudizievoli", None), ai_data.get("pregiudizievoli"))
        pregiudizievoli_dettaglio = stringify(ai_data.get("pregiudizievoli_dettaglio"))
        creditore_procedente = first_non_empty(
            getattr(asta, "creditore_procedente", None),
            ai_data.get("creditore_procedente"),
        )
        debiti_condominiali = stringify(ai_data.get("debiti_condominiali"))

        rischio_operazione = extract_risk(asta, ai_data)
        punti_attenzione = stringify(ai_data.get("punti_di_attenzione_investitore"))
        costi_probabili = stringify(ai_data.get("costi_probabili"))
        valutazione_operativa = stringify(ai_data.get("valutazione_operativa"))
        strategia_consigliata = stringify(ai_data.get("strategia_consigliata"))
        note_investitore = stringify(ai_data.get("note_investitore"))
        sintesi_finale = first_non_empty(getattr(asta, "sintesi", None), ai_data.get("sintesi"))

        stato_pratica = clean(getattr(asta, "stato_pratica", None))
        note_operative = clean(getattr(asta, "note_operativi", None))
        scheda_url = f"{base_url}/aste/{getattr(asta, 'id', '')}"

        row_analysis = [
            getattr(asta, "id", None),
            tribunale,
            rge,
            citta,
            indirizzo,
            lotto,
            data_asta,
            valore_perizia,
            prezzo_base,
            offerta_minima,
            descrizione_immobile,
            stato_manutentivo,
            catasto,
            foglio,
            mappale,
            subalterno,
            proprietario,
            tipologia_immobile,
            superficie,
            categoria_catastale,
            classe_catastale,
            rendita_catastale,
            agibilita,
            impianti,
            vincoli_oneri,
            conformita_catastale,
            stato_catastale,
            conformita_urbanistica,
            stato_urbanistico,
            abusi,
            criticita_principali,
            occupazione,
            stato_occupazione_dettaglio,
            pregiudizievoli,
            pregiudizievoli_dettaglio,
            creditore_procedente,
            debiti_condominiali,
            rischio_operazione,
            punti_attenzione,
            costi_probabili,
            valutazione_operativa,
            strategia_consigliata,
            note_investitore,
            sintesi_finale,
            stato_pratica,
            note_operative,
            "",
        ]

        style_data_row(ws2, row_idx, row_analysis)
        set_link(ws2.cell(row=row_idx, column=47), scheda_url, "Apri scheda")

        row_idx += 1

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions
    auto_fit_columns(ws2)

    # =========================================================
    # SALVATAGGIO
    # =========================================================
    if output_path is None:
        output_dir = Path("data")
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / "export_aste.xlsx"
    else:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

    wb.save(output_path)
    return str(output_path)